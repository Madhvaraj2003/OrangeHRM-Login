# Imported all the required modules
import os
import openpyxl
import pytest
from Page.login_page import LoginPage
 
# Function to read login test data from excel file
def read_login_data():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path =os.path.join(base_dir, '..', 'Test-data', 'login_data.xlsx') 
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active 
    data = []
    for row in sheet.iter_rows(min_row=2, values_only= True):
        if row[0] is not None:
            data.append((row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

# Collects each row's data (skips header)        
    return data


# Function to write test result (pass/fail) back into Excel file    
def write_result(test_id, result):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, '..', 'Test-data', 'login_data.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

# Find matching test ID and write  result in the "Test Result" column 
    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value).strip() == str(test_id).strip():
            sheet.cell(row=row[0].row, column=7).value = result
            break

# Save changes to workbook file 
    workbook.save(file_path)
    print(f"saved result '{result}' to excel")

# Parametrize the test function using the data ferom Excel
@pytest.mark.parametrize(
        "test_ID, username, password, test_result, date, time, tester, expected", 
        read_login_data()
        )  

# Function for testing login functionality 
def test_login_orm(browser, test_ID, username, password, test_result, date, time, tester, expected):
    page = LoginPage(browser)
    page.open_url()
    page.login(username, password)
    result = page.is_login_sucessful()

# Returns the test result
    test_result = "pass" if result else "fail" 
    print(f"TEST ID: {test_ID}, Username: {username}, Password: {password}")
    print(f"Test_result: {test_result}, Result: {result} ")
    print(f"Expected: {expected}")

    write_result(test_ID, test_result)
# Assertion to check if test result matches expected results from Excel     
    assert test_result == expected.lower()

    